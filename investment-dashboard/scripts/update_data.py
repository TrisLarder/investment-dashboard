import csv, io, json, os, statistics, time, zipfile, tempfile
from datetime import datetime, timezone
from pathlib import Path
import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data'/'latest.json'
HIST=ROOT/'data'/'history.json'
FRED_KEY=os.getenv('FRED_API_KEY','')
TWELVE_KEY=os.getenv('TWELVE_DATA_API_KEY','')
S=requests.Session(); S.headers.update({'User-Agent':'investment-dashboard/1.0'})

def pct(a,b): return None if a in (None,0) or b is None else (b/a-1)*100

BOE_YIELD_ZIP='https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip'
BUBA_SERIES={
    'de2':'D.REN.EUR.A610.000000WT0202.A',
    'de5':'D.REN.EUR.A620.000000WT0505.A',
    'de10':'D.REN.EUR.A630.000000WT1010.A',
}

def _to_float(x):
    try:
        if x is None or x == '': return None
        return float(str(x).replace(',','.'))
    except Exception:
        return None

def boe_nominal_spot():
    """Return daily BoE nominal gilt spot yields at 2Y/5Y/10Y.

    Uses the Bank's stable latest-yield-curve ZIP URL and parses the
    'GLC Nominal ... daily' workbook, standard 'spot curve' sheet.
    """
    r=S.get(BOE_YIELD_ZIP,timeout=60); r.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(r.content)) as z:
        names=z.namelist()
        cand=[n for n in names if 'GLC Nominal' in n and 'daily' in n.lower() and n.lower().endswith(('.xlsx','.xlsm'))]
        if not cand: raise RuntimeError('BoE nominal daily workbook not found in ZIP')
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
            f.write(z.read(cand[0])); f.flush()
            wb=load_workbook(f.name,read_only=True,data_only=True)
            sheets=[n for n in wb.sheetnames if n.strip().lower().endswith('spot curve')]
            if not sheets: raise RuntimeError('BoE spot curve sheet not found')
            ws=wb[sheets[0]]
            rows=list(ws.iter_rows(values_only=True))
    mat_idx=None; maturities=None
    for idx in [3,2,4,5]:  # Excel rows 4,3,5,6
        if idx >= len(rows): continue
        vals=[_to_float(v) for v in rows[idx][1:]]
        nums=[v for v in vals if v is not None and v>0]
        if len(nums)>=5 and all(nums[i] <= nums[i+1] for i in range(min(len(nums)-1,9))):
            mat_idx=idx; maturities=vals; break
    if mat_idx is None: raise RuntimeError('BoE maturity row not detected')
    targets={2.0:'uk2',5.0:'uk5',10.0:'uk10'}
    cols={}
    for i,m in enumerate(maturities, start=1):
        if m in targets: cols[targets[m]]=i
    if len(cols)<3: raise RuntimeError(f'BoE 2Y/5Y/10Y columns not found: {maturities[:25]}')
    out={k:[] for k in targets.values()}
    for row in rows[mat_idx+2:]:
        if not row: continue
        d=row[0]
        if hasattr(d,'date'): d=d.date().isoformat()
        elif hasattr(d,'isoformat'): d=d.isoformat()[:10]
        elif isinstance(d,(int,float)):
            from datetime import datetime, timedelta
            d=(datetime(1899,12,30)+timedelta(days=float(d))).date().isoformat()
        else:
            ds=str(d)[:10]
            if len(ds)!=10 or ds[4]!='-': continue
            d=ds
        for k,c in cols.items():
            if c < len(row):
                v=_to_float(row[c])
                if v is not None: out[k].append((d,v))
    return out

def bundesbank_current_yield(series_key):
    code=BUBA_SERIES[series_key]
    u=f'https://api.statistiken.bundesbank.de/rest/data/BBSSY/{code}'
    r=S.get(u,params={'format':'csv','lang':'en'},timeout=40); r.raise_for_status()
    text=r.text
    sample=text[:4096]
    try: dialect=csv.Sniffer().sniff(sample,delimiters=',;\t')
    except Exception: dialect=csv.excel
    rows=list(csv.DictReader(io.StringIO(text),dialect=dialect))
    vals=[]
    for x in rows:
        norm={str(k).strip().upper():v for k,v in x.items() if k is not None}
        d=norm.get('TIME_PERIOD') or norm.get('DATE') or norm.get('TIME PERIOD')
        v=norm.get('OBS_VALUE') or norm.get('VALUE') or norm.get('OBS VALUE')
        fv=_to_float(v)
        if d and fv is not None: vals.append((str(d)[:10],fv))
    vals.sort(key=lambda x:x[0])
    return vals[-10:]


def fred(series_id, n=10):
    if not FRED_KEY: return []
    u='https://api.stlouisfed.org/fred/series/observations'
    p={'series_id':series_id,'api_key':FRED_KEY,'file_type':'json','sort_order':'desc','limit':n}
    r=S.get(u,params=p,timeout=25); r.raise_for_status()
    out=[]
    for o in r.json().get('observations',[]):
        if o['value']!='.': out.append((o['date'],float(o['value'])))
    return out

def ecb_fx(ccy, n=220):
    # ECB EXR: units of currency per EUR, daily reference rate.
    u=f'https://data-api.ecb.europa.eu/service/data/EXR/D.{ccy}.EUR.SP00.A'
    r=S.get(u,params={'format':'csvdata','startPeriod':'2025-01-01'},timeout=30); r.raise_for_status()
    rows=list(csv.DictReader(io.StringIO(r.text)))
    vals=[(x['TIME_PERIOD'],float(x['OBS_VALUE'])) for x in rows if x.get('OBS_VALUE')]
    return vals[-n:]

def twelve(symbol, n=8):
    if not TWELVE_KEY: return []
    u='https://api.twelvedata.com/time_series'
    p={'symbol':symbol,'interval':'1day','outputsize':n,'apikey':TWELVE_KEY,'format':'JSON'}
    r=S.get(u,params=p,timeout=25); r.raise_for_status(); j=r.json()
    vals=[]
    for x in reversed(j.get('values',[])):
        vals.append((x['datetime'][:10],float(x['close'])))
    return vals

def put(series,key,label,kind,vals,source,prefix='',decimals=2,change_unit='%'):
    if not vals: return
    date,val=vals[-1]; prev=vals[-2][1] if len(vals)>1 else None; p5=vals[-6][1] if len(vals)>5 else None
    if kind in ('yield','spread'):
        c1=None if prev is None else (val-prev)*100
        c5=None if p5 is None else (val-p5)*100
        cu=' bp'
    else:
        c1=pct(prev,val); c5=pct(p5,val); cu='%'
    series[key]={'label':label,'kind':kind,'value':val,'date':date,'source':source,'prefix':prefix,'decimals':decimals,'change_1d':c1,'change_5d':c5,'change_unit':cu}

def main():
    base=json.loads(OUT.read_text()) if OUT.exists() else {'series':{},'sections':{},'headline':[],'signals':{}}
    s=base['series']
    mapping={'us2':('DGS2','US Treasury 2Y'),'us5':('DGS5','US Treasury 5Y'),'us10':('DGS10','US Treasury 10Y')}
    for k,(fid,lbl) in mapping.items(): put(s,k,lbl,'yield',list(reversed(fred(fid,10))),'FRED / U.S. Treasury')

    # Official UK gilt curve from Bank of England; published with a short delay.
    try:
        uk=boe_nominal_spot()
        for k,lbl in [('uk2','UK Gilt 2Y'),('uk5','UK Gilt 5Y'),('uk10','UK Gilt 10Y')]:
            put(s,k,lbl,'yield',uk.get(k,[]),'Bank of England fitted nominal gilt spot curve')
    except Exception as e:
        print('BoE yields',e)

    # Official German benchmark issues from Deutsche Bundesbank.
    for k,lbl in [('de2','Germany 2Y'),('de5','Germany 5Y'),('de10','Germany 10Y')]:
        try: put(s,k,lbl,'yield',bundesbank_current_yield(k),'Deutsche Bundesbank current Federal securities')
        except Exception as e: print('Bundesbank',k,e)
    put(s,'wti','WTI Crude','price',list(reversed(fred('DCOILWTICO',10))),'EIA via FRED',prefix='$')
    put(s,'brent','Brent Crude','price',list(reversed(fred('DCOILBRENTEU',10))),'EIA via FRED',prefix='$')
    put(s,'vix','VIX','index',list(reversed(fred('VIXCLS',10))),'CBOE via FRED')
    put(s,'hy_oas','US HY OAS','spread',list(reversed(fred('BAMLH0A0HYM2',10))),'ICE BofA via FRED')

    fx={}
    for c in ['USD','GBP','CAD','JPY','AUD']:
        try: fx[c]=ecb_fx(c)
        except Exception as e: print('ECB',c,e)
    if all(c in fx and fx[c] for c in ['USD','GBP']):
        dates=sorted(set(dict(fx['USD'])) & set(dict(fx['GBP'])))
        vals=[(d,dict(fx['USD'])[d]/dict(fx['GBP'])[d]) for d in dates]; put(s,'gbpusd','GBP / USD','fx',vals,'ECB',decimals=4)
    if all(c in fx and fx[c] for c in ['CAD','GBP']):
        dates=sorted(set(dict(fx['CAD'])) & set(dict(fx['GBP'])))
        vals=[(d,dict(fx['CAD'])[d]/dict(fx['GBP'])[d]) for d in dates]; put(s,'gbpcad','GBP / CAD','fx',vals,'ECB',decimals=4)
    if all(c in fx and fx[c] for c in ['JPY','USD']):
        dates=sorted(set(dict(fx['JPY'])) & set(dict(fx['USD'])))
        vals=[(d,dict(fx['JPY'])[d]/dict(fx['USD'])[d]) for d in dates]; put(s,'usdjpy','USD / JPY','fx',vals,'ECB',decimals=2)
        recent=vals[-200:] if len(vals)>=200 else vals
        ma=statistics.mean(v for _,v in recent) if recent else None
        if ma:
            cur=vals[-1][1]; s['carry_200d']={'label':'USD/JPY vs 200D','kind':'ratio','value':(cur/ma-1)*100,'suffix':'%','date':vals[-1][0],'source':'derived from ECB','decimals':2,'signal':'ARM' if cur<ma else 'NORMAL'}
    if all(c in fx and fx[c] for c in ['JPY','AUD']):
        dates=sorted(set(dict(fx['JPY'])) & set(dict(fx['AUD'])))
        vals=[(d,dict(fx['JPY'])[d]/dict(fx['AUD'])[d]) for d in dates]; put(s,'audjpy','AUD / JPY','fx',vals,'ECB',decimals=2)

    for key,sym,label in [('gold','XAU/USD','Gold'),('silver','XAG/USD','Silver'),('platinum','XPT/USD','Platinum')]:
        try: put(s,key,label,'price',twelve(sym), 'Twelve Data',prefix='$')
        except Exception as e: print('Twelve',sym,e)


    if s.get('gold',{}).get('value') and s.get('silver',{}).get('value'):
        g=s['gold']; q=s['silver']; s['gold_silver']={'label':'Gold / Silver Ratio','kind':'ratio','value':g['value']/q['value'],'date':min(g['date'],q['date']),'source':'derived','decimals':1}
    if s.get('brent',{}).get('value') is not None and s.get('wti',{}).get('value') is not None:
        b,w=s['brent'],s['wti']; s['brent_wti']={'label':'Brent – WTI Spread','kind':'price','prefix':'$','value':b['value']-w['value'],'date':min(b['date'],w['date']),'source':'derived','decimals':2}

    carry_status='NORMAL'; active=[]; high=False
    uj=s.get('usdjpy',{}); cv=s.get('carry_200d',{})
    if cv.get('value') is not None and cv['value']<0: active.append('USD/JPY below 200D')
    yen_move=None
    if uj.get('change_1d') is not None: yen_move=-uj['change_1d']
    if yen_move is not None and yen_move>2: active.append('Yen >2%'); high=True
    hy=s.get('hy_oas',{}); hy5=hy.get('change_5d')
    if hy5 is not None and hy5>=75: active.append('HY +75bp/5D'); high=True
    vix=s.get('vix',{}).get('value')
    if len(active)>=3 and high: carry_status='FIRE'
    elif high: carry_status='CONFIRM'
    elif active: carry_status='ARM'
    base['signals']['carry']={'status':carry_status,'summary':('No high-specificity forced-deleveraging confirmation.' if not high else 'High-specificity stress confirmation is active.'),'drivers':[{'label':'USD/JPY vs 200D','value':('—' if cv.get('value') is None else f"{cv['value']:+.2f}%")},{'label':'Yen 1D','value':('—' if yen_move is None else f'{yen_move:+.2f}%')},{'label':'HY 5D','value':('—' if hy5 is None else f'{hy5:+.0f} bp')} ]}

    # Real estate funding pulse: directional, transparent, not valuation advice.
    drivers=[]; score=0
    for k in ['uk5','de5','hy_oas']:
        x=s.get(k,{}); c=x.get('change_5d');
        if c is not None: score += -1 if c>0 else 1 if c<0 else 0
    re_status='SUPPORTIVE' if score>=2 else 'TIGHTENING' if score<=-2 else 'NEUTRAL'
    base['signals']['real_estate']={'status':re_status,'summary':'Directional funding pulse based on 5-day moves in medium-term sovereign yields and credit spreads; it is not a property valuation model.','drivers':[{'label':'UK 5Y','value':('—' if s.get('uk5',{}).get('change_5d') is None else f"{s['uk5']['change_5d']:+.0f} bp")},{'label':'Germany 5Y','value':('—' if s.get('de5',{}).get('change_5d') is None else f"{s['de5']['change_5d']:+.0f} bp")},{'label':'HY OAS','value':('—' if hy5 is None else f'{hy5:+.0f} bp')} ]}

    base['generated_at']=datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')
    base['market_status']='Latest available'; base['build']=os.getenv('GITHUB_SHA','local')[:7]
    OUT.write_text(json.dumps(base,indent=2))
    hist=[]
    if HIST.exists():
        try: hist=json.loads(HIST.read_text())
        except: pass
    hist.append({'generated_at':base['generated_at'],'series':{k:{'value':v.get('value'),'date':v.get('date')} for k,v in s.items()}})
    HIST.write_text(json.dumps(hist[-370:],indent=2))

if __name__=='__main__': main()
