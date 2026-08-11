"""Fetch and normalise the dashboard's delayed/public market data.

Every provider is isolated: a failed request retains the last valid observation
and records feed health rather than blanking the dashboard or aborting the run.
"""

import csv
import io
import json
import os
import statistics
import tempfile
import zipfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "latest.json"
HIST = ROOT / "data" / "history.json"
FRED_KEY = os.getenv("FRED_API_KEY", "").strip()
TWELVE_KEY = os.getenv("TWELVE_DATA_API_KEY", "").strip()
S = requests.Session()
S.headers.update({"User-Agent": "investment-dashboard/2.0 (+GitHub Pages)"})

BOE_YIELD_ZIP = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip"
BOE_IADB = "https://www.bankofengland.co.uk/boeapps/database/_iadb-fromshowcolumns.asp"
JGB_CSV = "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/historical/jgbcme_all.csv"
CFTC_TFF_ZIP = "https://www.cftc.gov/files/dea/history/fut_fin_txt_{year}.zip"
CBOE_SETTLEMENT = "https://www-api.cboe.com/us/futures/market_statistics/settlement/csv"
BUBA_SERIES = {
    "de2": "D.REN.EUR.A610.000000WT0202.A",
    "de5": "D.REN.EUR.A620.000000WT0505.A",
    "de10": "D.REN.EUR.A630.000000WT1010.A",
}


def pct(old, new):
    return None if old in (None, 0) or new is None else (new / old - 1) * 100


def _to_float(value):
    try:
        if value is None or str(value).strip() in ("", ".", "-"):
            return None
        return float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def _clean_error(exc):
    text = " ".join(str(exc).split())
    return text[:220] or exc.__class__.__name__


def _date_from_timestamp(ts):
    return datetime.fromtimestamp(ts, tz=timezone.utc).date().isoformat()


def _sorted_values(values):
    return sorted({str(d)[:10]: float(v) for d, v in values if v is not None}.items())


def boe_nominal_spot():
    response = S.get(BOE_YIELD_ZIP, timeout=60)
    response.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        candidates = [
            name
            for name in archive.namelist()
            if "GLC Nominal" in name
            and "daily" in name.lower()
            and name.lower().endswith((".xlsx", ".xlsm"))
        ]
        if not candidates:
            raise RuntimeError("BoE nominal daily workbook not found")
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp:
            temp.write(archive.read(candidates[0]))
            temp.flush()
            workbook = load_workbook(temp.name, read_only=True, data_only=True)
            sheets = [n for n in workbook.sheetnames if n.strip().lower().endswith("spot curve")]
            if not sheets:
                raise RuntimeError("BoE spot curve sheet not found")
            rows = list(workbook[sheets[0]].iter_rows(values_only=True))

    maturity_row = maturities = None
    for index in (3, 2, 4, 5):
        if index >= len(rows):
            continue
        values = [_to_float(value) for value in rows[index][1:]]
        numbers = [value for value in values if value is not None and value > 0]
        if len(numbers) >= 5 and all(numbers[i] <= numbers[i + 1] for i in range(min(len(numbers) - 1, 9))):
            maturity_row, maturities = index, values
            break
    if maturity_row is None:
        raise RuntimeError("BoE maturity row not detected")

    target = {2.0: "uk2", 5.0: "uk5", 10.0: "uk10"}
    columns = {target[m]: i for i, m in enumerate(maturities, start=1) if m in target}
    if len(columns) != 3:
        raise RuntimeError("BoE 2Y/5Y/10Y columns not found")
    result = {key: [] for key in target.values()}
    for row in rows[maturity_row + 2 :]:
        if not row:
            continue
        raw_date = row[0]
        if hasattr(raw_date, "date"):
            day = raw_date.date().isoformat()
        elif isinstance(raw_date, (int, float)):
            day = (datetime(1899, 12, 30) + timedelta(days=float(raw_date))).date().isoformat()
        else:
            day = str(raw_date)[:10]
            if len(day) != 10 or day[4] != "-":
                continue
        for key, column in columns.items():
            value = _to_float(row[column]) if column < len(row) else None
            if value is not None:
                result[key].append((day, value))
    return result


def bundesbank_current_yield(series_key):
    """Read official SDMX-CSV. The Accept header is required; format=csv is not."""
    url = f"https://api.statistiken.bundesbank.de/rest/data/BBSSY/{BUBA_SERIES[series_key]}"
    response = S.get(
        url,
        params={"lastNObservations": 12, "detail": "dataonly"},
        headers={"Accept": "text/csv", "Accept-Language": "en"},
        timeout=40,
    )
    response.raise_for_status()
    rows = csv.DictReader(io.StringIO(response.text), delimiter=";")
    values = []
    for row in rows:
        day = row.get("TIME_PERIOD")
        value = _to_float(row.get("OBS_VALUE"))
        if day and value is not None:
            values.append((day[:10], value))
    if not values:
        raise RuntimeError("Bundesbank returned no observations")
    return _sorted_values(values)[-10:]


def jgb_constant_maturity():
    response = S.get(JGB_CSV, timeout=75)
    response.raise_for_status()
    lines = response.content.decode("utf-8-sig").splitlines()
    if lines and lines[0].startswith("Interest Rate"):
        lines = lines[1:]
    rows = list(csv.DictReader(lines, skipinitialspace=True))
    result = {"jp2": [], "jp5": [], "jp10": []}
    for row in rows:
        raw_day = (row.get("Date") or "").strip()
        if not raw_day:
            continue
        try:
            day = datetime.strptime(raw_day, "%Y/%m/%d").date().isoformat()
        except ValueError:
            continue
        for key, column in (("jp2", "2Y"), ("jp5", "5Y"), ("jp10", "10Y")):
            value = _to_float(row.get(column))
            if value is not None:
                result[key].append((day, value))
    if not all(result.values()):
        raise RuntimeError("Japan MOF CSV missing 2Y/5Y/10Y observations")
    return {key: values[-10:] for key, values in result.items()}


def fred(series_id, count=10):
    if not FRED_KEY:
        raise RuntimeError("FRED_API_KEY is not configured")
    response = S.get(
        "https://api.stlouisfed.org/fred/series/observations",
        params={
            "series_id": series_id,
            "api_key": FRED_KEY,
            "file_type": "json",
            "sort_order": "desc",
            "limit": count,
        },
        timeout=25,
    )
    response.raise_for_status()
    values = [(item["date"], _to_float(item["value"])) for item in response.json().get("observations", [])]
    values = [(day, value) for day, value in values if value is not None]
    if not values:
        raise RuntimeError(f"FRED {series_id} returned no observations")
    return list(reversed(values))


def ecb_series(flow, key, count=12):
    response = S.get(
        f"https://data-api.ecb.europa.eu/service/data/{flow}/{key}",
        params={"format": "csvdata", "lastNObservations": count},
        timeout=35,
    )
    response.raise_for_status()
    rows = csv.DictReader(io.StringIO(response.text))
    values = []
    for row in rows:
        value = _to_float(row.get("OBS_VALUE"))
        if row.get("TIME_PERIOD") and value is not None:
            values.append((row["TIME_PERIOD"][:10], value))
    if not values:
        raise RuntimeError(f"ECB {flow} returned no observations")
    return _sorted_values(values)


def ecb_fx(currency, count=260):
    return ecb_series("EXR", f"D.{currency}.EUR.SP00.A", count)


def boe_iadb(series_code, count=12):
    today = date.today()
    start = today - timedelta(days=45)
    params = {
        "csv.x": "yes",
        "Datefrom": start.strftime("%d/%b/%Y"),
        "Dateto": today.strftime("%d/%b/%Y"),
        "SeriesCodes": series_code,
        "CSVF": "TN",
        "UsingCodes": "Y",
        "VPD": "Y",
        "VFD": "N",
    }
    response = S.get(BOE_IADB, params=params, timeout=40)
    response.raise_for_status()
    rows = list(csv.reader(io.StringIO(response.text)))
    values = []
    for row in rows:
        if len(row) < 2:
            continue
        value = _to_float(row[1])
        if value is None:
            continue
        parsed = None
        for pattern in ("%d %b %Y", "%d-%b-%Y", "%d/%m/%Y"):
            try:
                parsed = datetime.strptime(row[0].strip(), pattern).date().isoformat()
                break
            except ValueError:
                pass
        if parsed:
            values.append((parsed, value))
    if not values:
        raise RuntimeError(f"BoE {series_code} returned no observations")
    return _sorted_values(values)[-count:]


def twelve(symbol, count=8):
    if not TWELVE_KEY:
        raise RuntimeError("TWELVE_DATA_API_KEY is not configured")
    response = S.get(
        "https://api.twelvedata.com/time_series",
        params={"symbol": symbol, "interval": "1day", "outputsize": count, "apikey": TWELVE_KEY, "format": "JSON"},
        timeout=25,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("status") == "error" or payload.get("code"):
        raise RuntimeError(f"Twelve Data {symbol}: {payload.get('message', 'API error')}")
    values = [(item.get("datetime", "")[:10], _to_float(item.get("close"))) for item in payload.get("values", [])]
    values = [(day, value) for day, value in values if day and value is not None]
    if not values:
        raise RuntimeError(f"Twelve Data {symbol} returned no prices")
    return _sorted_values(values)


def yahoo_chart(symbol, count=10):
    response = S.get(
        f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}",
        params={"range": "1mo", "interval": "1d", "events": "history"},
        headers={"User-Agent": "Mozilla/5.0 investment-dashboard/2.0"},
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json().get("chart", {})
    if payload.get("error"):
        raise RuntimeError(payload["error"].get("description", "Yahoo chart error"))
    result = (payload.get("result") or [None])[0]
    if not result:
        raise RuntimeError(f"Yahoo {symbol} returned no result")
    timestamps = result.get("timestamp") or []
    closes = (((result.get("indicators") or {}).get("quote") or [{}])[0].get("close") or [])
    values = [(_date_from_timestamp(ts), _to_float(value)) for ts, value in zip(timestamps, closes)]
    values = [(day, value) for day, value in values if value is not None]
    if not values:
        raise RuntimeError(f"Yahoo {symbol} returned no prices")
    return _sorted_values(values)[-count:]


def cftc_yen_positioning():
    response = S.get(CFTC_TFF_ZIP.format(year=date.today().year), timeout=60)
    response.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        filename = next((name for name in archive.namelist() if name.lower().endswith(".txt")), None)
        if not filename:
            raise RuntimeError("CFTC TFF archive has no text file")
        text = archive.read(filename).decode("utf-8-sig", errors="replace")
    rows = csv.DictReader(io.StringIO(text), skipinitialspace=True)
    values = []
    for row in rows:
        if "JAPANESE YEN" not in (row.get("Market_and_Exchange_Names") or "").upper():
            continue
        long_value = _to_float(row.get("Lev_Money_Positions_Long_All"))
        short_value = _to_float(row.get("Lev_Money_Positions_Short_All"))
        open_interest = _to_float(row.get("Open_Interest_All"))
        if long_value is None or short_value is None:
            continue
        net = long_value - short_value
        values.append(
            {
                "date": row.get("Report_Date_as_YYYY-MM-DD", "")[:10],
                "value": net,
                "long": long_value,
                "short": short_value,
                "open_interest": open_interest,
                "pct_oi": None if not open_interest else net / open_interest * 100,
            }
        )
    if not values:
        raise RuntimeError("CFTC TFF archive has no Japanese yen rows")
    return sorted(values, key=lambda item: item["date"])


def cboe_vix_curve():
    """Return the latest two distinct standard VX monthly settlements."""
    for days_back in range(0, 9):
        day = date.today() - timedelta(days=days_back)
        if day.weekday() >= 5:
            continue
        response = S.get(CBOE_SETTLEMENT, params={"dt": day.isoformat()}, timeout=30)
        if response.status_code != 200:
            continue
        rows = list(csv.DictReader(io.StringIO(response.text)))
        contracts = []
        seen_expiries = set()
        for row in rows:
            symbol = (row.get("Symbol") or "").strip()
            expiry = (row.get("Expiration Date") or "").strip()
            price = _to_float(row.get("Price"))
            # Standard monthly contracts are VX/<month code>; weeklys include digits.
            stem = symbol.split("/")[0]
            if row.get("Product") != "VX" or stem != "VX" or not expiry or price is None or expiry in seen_expiries:
                continue
            seen_expiries.add(expiry)
            contracts.append((expiry, price, symbol))
        contracts.sort()
        if len(contracts) >= 2:
            return {"date": day.isoformat(), "front": contracts[0], "second": contracts[1]}
    raise RuntimeError("Cboe returned no two-month standard VX curve")


def put(series, key, label, kind, values, source, prefix="", decimals=2, stale_days=None, extra=None):
    values = _sorted_values(values)
    if not values:
        raise RuntimeError(f"{label} returned no observations")
    day, value = values[-1]
    prior = values[-2][1] if len(values) > 1 else None
    five_back = values[-6][1] if len(values) > 5 else None
    if kind in ("yield", "spread"):
        change_1d = None if prior is None else (value - prior) * 100
        change_5d = None if five_back is None else (value - five_back) * 100
        change_unit = " bp"
    else:
        change_1d, change_5d, change_unit = pct(prior, value), pct(five_back, value), "%"
    item = {
        "label": label,
        "kind": kind,
        "value": value,
        "date": day,
        "source": source,
        "prefix": prefix,
        "decimals": decimals,
        "change_1d": change_1d,
        "change_5d": change_5d,
        "change_unit": change_unit,
        "stale": False,
        "feed_status": "ok",
    }
    if stale_days:
        try:
            item["stale"] = (date.today() - date.fromisoformat(day)).days > stale_days
        except ValueError:
            pass
    if extra:
        item.update(extra)
    series[key] = item


def unavailable(series, key, label, kind, source, message, prefix="", decimals=2):
    previous = series.get(key, {})
    if previous.get("value") is not None:
        previous.update({"stale": True, "feed_status": "stale", "feed_error": message})
        series[key] = previous
    else:
        series[key] = {
            "label": label,
            "kind": kind,
            "value": None,
            "date": "Unavailable",
            "source": source,
            "prefix": prefix,
            "decimals": decimals,
            "stale": True,
            "feed_status": "unavailable",
            "feed_error": message,
        }


def add_feed(feed_health, name, action):
    try:
        action()
        feed_health.append({"name": name, "status": "ok"})
        return True
    except Exception as exc:  # provider failures must not abort unrelated feeds
        message = _clean_error(exc)
        print(f"{name}: {message}")
        feed_health.append({"name": name, "status": "error", "message": message})
        return False


def derived(series, key, label, kind, value, day, source="Derived", **kwargs):
    series[key] = {
        "label": label,
        "kind": kind,
        "value": value,
        "date": day,
        "source": source,
        "decimals": kwargs.pop("decimals", 2),
        "stale": kwargs.pop("stale", False),
        "feed_status": "derived",
        **kwargs,
    }


def fetch_with_fallback(series, feed_health, key, label, symbol, fallback_symbol, fallback_source, kind="price", prefix="$", decimals=2):
    errors = []
    if TWELVE_KEY:
        try:
            put(series, key, label, kind, twelve(symbol), "Twelve Data", prefix=prefix, decimals=decimals, stale_days=4)
            feed_health.append({"name": f"Twelve Data {symbol}", "status": "ok"})
            return
        except Exception as exc:
            errors.append(_clean_error(exc))
            feed_health.append({"name": f"Twelve Data {symbol}", "status": "error", "message": errors[-1]})
    else:
        errors.append("TWELVE_DATA_API_KEY is not configured")
    try:
        put(series, key, f"{label} (futures proxy)", kind, yahoo_chart(fallback_symbol), fallback_source, prefix=prefix, decimals=decimals, stale_days=4, extra={"proxy": True})
        feed_health.append({"name": fallback_source, "status": "fallback" if TWELVE_KEY else "ok"})
    except Exception as exc:
        errors.append(_clean_error(exc))
        unavailable(series, key, label, kind, fallback_source, " | ".join(errors), prefix=prefix, decimals=decimals)
        feed_health.append({"name": fallback_source, "status": "error", "message": errors[-1]})


def main():
    base = json.loads(OUT.read_text(encoding="utf-8")) if OUT.exists() else {"series": {}, "signals": {}}
    series = base.setdefault("series", {})
    feed_health = []

    for key, fred_id, label in (
        ("us2", "DGS2", "US Treasury 2Y"),
        ("us5", "DGS5", "US Treasury 5Y"),
        ("us10", "DGS10", "US Treasury 10Y"),
        ("wti", "DCOILWTICO", "WTI Crude"),
        ("brent", "DCOILBRENTEU", "Brent Crude"),
        ("vix", "VIXCLS", "VIX"),
        ("hy_oas", "BAMLH0A0HYM2", "US HY OAS"),
    ):
        kind = "yield" if key.startswith("us") else "spread" if key == "hy_oas" else "index" if key == "vix" else "price"
        prefix = "$" if key in ("wti", "brent") else ""
        source = "FRED / U.S. Treasury" if key.startswith("us") else "EIA via FRED" if key in ("wti", "brent") else "Cboe via FRED" if key == "vix" else "ICE BofA via FRED"
        ok = add_feed(feed_health, f"FRED {fred_id}", lambda k=key, f=fred_id, l=label, kd=kind, src=source, pre=prefix: put(series, k, l, kd, fred(f), src, prefix=pre, stale_days=8))
        if not ok:
            unavailable(series, key, label, kind, source, feed_health[-1]["message"], prefix=prefix)

    def uk_action():
        values = boe_nominal_spot()
        for key, label in (("uk2", "UK Gilt 2Y"), ("uk5", "UK Gilt 5Y"), ("uk10", "UK Gilt 10Y")):
            put(series, key, label, "yield", values[key], "Bank of England fitted nominal gilt spot curve", stale_days=8)

    if not add_feed(feed_health, "Bank of England gilt curve", uk_action):
        for key, label in (("uk2", "UK Gilt 2Y"), ("uk5", "UK Gilt 5Y"), ("uk10", "UK Gilt 10Y")):
            unavailable(series, key, label, "yield", "Bank of England", feed_health[-1]["message"])

    for key, label in (("de2", "Germany 2Y"), ("de5", "Germany 5Y"), ("de10", "Germany 10Y")):
        ok = add_feed(feed_health, f"Bundesbank {label}", lambda k=key, l=label: put(series, k, l, "yield", bundesbank_current_yield(k), "Deutsche Bundesbank current Federal securities", stale_days=8))
        if not ok:
            unavailable(series, key, label, "yield", "Deutsche Bundesbank", feed_health[-1]["message"])

    def jgb_action():
        values = jgb_constant_maturity()
        for key, label in (("jp2", "Japan JGB 2Y"), ("jp5", "Japan JGB 5Y"), ("jp10", "Japan JGB 10Y")):
            put(series, key, label, "yield", values[key], "Japan Ministry of Finance constant-maturity curve", stale_days=14)

    if not add_feed(feed_health, "Japan MOF JGB curve", jgb_action):
        for key, label in (("jp2", "Japan JGB 2Y"), ("jp5", "Japan JGB 5Y"), ("jp10", "Japan JGB 10Y")):
            unavailable(series, key, label, "yield", "Japan Ministry of Finance", feed_health[-1]["message"])

    fx = {}
    for currency in ("USD", "GBP", "CAD", "JPY", "AUD"):
        add_feed(feed_health, f"ECB FX {currency}", lambda c=currency: fx.update({c: ecb_fx(c)}))

    def cross(left, right, key, label, decimals):
        if left not in fx or right not in fx:
            unavailable(series, key, label, "fx", "ECB reference rates", "Required ECB leg unavailable", decimals=decimals)
            return []
        left_map, right_map = dict(fx[left]), dict(fx[right])
        values = [(day, left_map[day] / right_map[day]) for day in sorted(set(left_map) & set(right_map))]
        put(series, key, label, "fx", values, "ECB reference rates", decimals=decimals, stale_days=5)
        return values

    gbpusd = cross("USD", "GBP", "gbpusd", "GBP / USD", 4)
    cross("CAD", "GBP", "gbpcad", "GBP / CAD", 4)
    usdjpy = cross("JPY", "USD", "usdjpy", "USD / JPY", 2)
    audjpy = cross("JPY", "AUD", "audjpy", "AUD / JPY", 2)

    for key, values, label in (("carry_usdjpy_200d", usdjpy, "USD/JPY vs 200D"), ("carry_audjpy_200d", audjpy, "AUD/JPY vs 200D")):
        if values:
            window = values[-200:]
            average = statistics.mean(value for _, value in window)
            distance = (values[-1][1] / average - 1) * 100
            derived(series, key, label, "ratio", distance, values[-1][0], "Derived from ECB reference rates", suffix="%", decimals=2, signal="ARM" if distance < 0 else "NORMAL")

    fetch_with_fallback(series, feed_health, "gold", "Gold spot", "XAU/USD", "GC=F", "COMEX gold futures fallback via Yahoo Finance")
    fetch_with_fallback(series, feed_health, "silver", "Silver spot", "XAG/USD", "SI=F", "COMEX silver futures fallback via Yahoo Finance")
    fetch_with_fallback(series, feed_health, "platinum", "Platinum spot", "XPT/USD", "PL=F", "NYMEX platinum futures fallback via Yahoo Finance")

    for key, label, symbol, source, prefix in (
        ("copper", "Copper (COMEX)", "HG=F", "COMEX copper futures via Yahoo Finance", "$"),
        ("uranium", "Uranium equities proxy (URA)", "URA", "Global X Uranium ETF via Yahoo Finance", "$"),
    ):
        ok = add_feed(feed_health, source, lambda k=key, l=label, sy=symbol, src=source, pre=prefix: put(series, k, l, "price", yahoo_chart(sy), src, prefix=pre, stale_days=4, extra={"proxy": key == "uranium"}))
        if not ok:
            unavailable(series, key, label, "price", source, feed_health[-1]["message"], prefix=prefix)

    def cftc_action():
        observations = cftc_yen_positioning()
        values = [(item["date"], item["value"]) for item in observations]
        latest = observations[-1]
        put(series, "cftc_yen", "CFTC yen leveraged funds", "position", values, "CFTC Traders in Financial Futures", decimals=0, stale_days=12, extra={"long": latest["long"], "short": latest["short"], "pct_oi": latest["pct_oi"]})

    if not add_feed(feed_health, "CFTC yen TFF", cftc_action):
        unavailable(series, "cftc_yen", "CFTC yen leveraged funds", "position", "CFTC Traders in Financial Futures", feed_health[-1]["message"])

    def vix_curve_action():
        curve = cboe_vix_curve()
        front, second = curve["front"], curve["second"]
        derived(series, "vix_front", "VIX front future", "index", front[1], curve["date"], "Cboe official settlement", decimals=2, expiry=front[0])
        derived(series, "vix_second", "VIX second future", "index", second[1], curve["date"], "Cboe official settlement", decimals=2, expiry=second[0])
        spread = front[1] - second[1]
        derived(series, "vix_curve", "VIX front – second", "index", spread, curve["date"], "Derived from Cboe official settlements", decimals=2, signal="BACKWARDATION" if spread > 0 else "CONTANGO")

    if not add_feed(feed_health, "Cboe VIX settlements", vix_curve_action):
        unavailable(series, "vix_curve", "VIX front – second", "index", "Cboe official settlement", feed_health[-1]["message"])

    funding_fetches = (
        ("sonia", "SONIA", "yield", lambda: boe_iadb("IUDSOIA"), "Bank of England", 8),
        ("estr", "€STR", "yield", lambda: ecb_series("EST", "B.EU000A2X2A25.WT", 12), "European Central Bank", 8),
        ("ecb_deposit", "ECB deposit facility", "yield", lambda: ecb_series("FM", "D.U2.EUR.4F.KR.DFR.LEV", 12), "European Central Bank", 45),
    )
    for key, label, kind, getter, source, stale_days in funding_fetches:
        ok = add_feed(feed_health, label, lambda k=key, l=label, kd=kind, g=getter, src=source, sd=stale_days: put(series, k, l, kd, g(), src, stale_days=sd))
        if not ok:
            unavailable(series, key, label, kind, source, feed_health[-1]["message"])

    if series.get("us2", {}).get("value") is not None and series.get("jp2", {}).get("value") is not None:
        day = min(series["us2"]["date"], series["jp2"]["date"])
        derived(series, "us_jp_2y", "US–Japan 2Y spread", "spread", (series["us2"]["value"] - series["jp2"]["value"]) * 100, day, decimals=0, suffix=" bp")
    if series.get("gold", {}).get("value") and series.get("silver", {}).get("value"):
        day = min(series["gold"]["date"], series["silver"]["date"])
        derived(series, "gold_silver", "Gold / Silver ratio", "ratio", series["gold"]["value"] / series["silver"]["value"], day, decimals=1)
    if series.get("copper", {}).get("value") and series.get("gold", {}).get("value"):
        day = min(series["copper"]["date"], series["gold"]["date"])
        derived(series, "copper_gold", "Copper / Gold", "ratio", series["copper"]["value"] / series["gold"]["value"] * 1000, day, decimals=2)
    if series.get("brent", {}).get("value") is not None and series.get("wti", {}).get("value") is not None:
        day = min(series["brent"]["date"], series["wti"]["date"])
        derived(series, "brent_wti", "Brent – WTI spread", "price", series["brent"]["value"] - series["wti"]["value"], day, prefix="$", decimals=2)

    hy = series.get("hy_oas", {})
    for key, label, sovereign in (("uk_funding_proxy", "UK 5Y funding proxy", "uk5"), ("eu_funding_proxy", "Euro 5Y funding proxy", "de5")):
        if series.get(sovereign, {}).get("value") is not None and hy.get("value") is not None:
            day = min(series[sovereign]["date"], hy["date"])
            derived(series, key, label, "yield", series[sovereign]["value"] + hy["value"], day, "Derived: sovereign 5Y + US HY OAS (directional, not a lending quote)", decimals=2)

    us_distance = series.get("carry_usdjpy_200d", {}).get("value")
    au_distance = series.get("carry_audjpy_200d", {}).get("value")
    yen_move = None if series.get("usdjpy", {}).get("change_1d") is None else -series["usdjpy"]["change_1d"]
    hy_change = hy.get("change_5d")
    vix_spread = series.get("vix_curve", {}).get("value")
    cftc_pct = series.get("cftc_yen", {}).get("pct_oi")
    active = []
    high_specificity = False
    if us_distance is not None and us_distance < 0:
        active.append("USD/JPY below 200D")
    if au_distance is not None and au_distance < 0:
        active.append("AUD/JPY below 200D")
    if yen_move is not None and yen_move > 2:
        active.append("Yen appreciation >2%")
        high_specificity = True
    if vix_spread is not None and vix_spread > 0:
        active.append("VIX backwardation")
        high_specificity = True
    if hy_change is not None and hy_change >= 75:
        active.append("HY OAS +75bp/5D")
        high_specificity = True
    if cftc_pct is not None and cftc_pct <= -12:
        active.append("Crowded leveraged yen short")
    carry_status = "FIRE" if len(active) >= 3 and high_specificity else "CONFIRM" if high_specificity else "ARM" if active else "NORMAL"
    base.setdefault("signals", {})["carry"] = {
        "status": carry_status,
        "summary": "No high-specificity forced-deleveraging confirmation." if not high_specificity else "At least one high-specificity stress confirmation is active.",
        "drivers": [
            {"label": "USD/JPY vs 200D", "value": "—" if us_distance is None else f"{us_distance:+.2f}%"},
            {"label": "AUD/JPY vs 200D", "value": "—" if au_distance is None else f"{au_distance:+.2f}%"},
            {"label": "VIX curve", "value": "—" if vix_spread is None else ("Backwardation" if vix_spread > 0 else "Contango")},
            {"label": "CFTC yen", "value": "—" if cftc_pct is None else f"{cftc_pct:+.1f}% OI"},
        ],
    }

    drivers, score = [], 0
    for key, label in (("uk5", "UK 5Y"), ("de5", "Germany 5Y"), ("hy_oas", "HY OAS")):
        change = series.get(key, {}).get("change_5d")
        if change is not None:
            score += -1 if change > 0 else 1 if change < 0 else 0
        drivers.append({"label": label, "value": "—" if change is None else f"{change:+.0f} bp"})
    for key, label in (("sonia", "SONIA"), ("estr", "€STR")):
        value = series.get(key, {}).get("value")
        drivers.append({"label": label, "value": "—" if value is None else f"{value:.2f}%"})
    real_estate_status = "SUPPORTIVE" if score >= 2 else "TIGHTENING" if score <= -2 else "NEUTRAL"
    base["signals"]["real_estate"] = {
        "status": real_estate_status,
        "summary": "Directional funding pulse from medium-term sovereign yields, overnight benchmarks and credit spreads; derived proxies are not executable loan quotes.",
        "drivers": drivers,
    }

    base["headline"] = ["uk10", "de10", "us10", "jp10", "gbpusd", "usdjpy", "gold", "silver", "copper", "uranium", "hy_oas", "vix"]
    base["sections"] = {
        "rates": ["uk2", "uk5", "uk10", "de2", "de5", "de10", "us2", "us5", "us10", "jp2", "jp5", "jp10", "us_jp_2y"],
        "fx": ["gbpusd", "gbpcad", "usdjpy", "audjpy", "carry_usdjpy_200d", "carry_audjpy_200d", "cftc_yen"],
        "commodities": ["gold", "silver", "platinum", "copper", "uranium", "wti", "brent", "gold_silver", "copper_gold", "brent_wti"],
        "realestate": ["sonia", "estr", "ecb_deposit", "uk5", "de5", "hy_oas", "uk_funding_proxy", "eu_funding_proxy"],
        "risk": ["vix", "vix_front", "vix_second", "vix_curve", "hy_oas", "cftc_yen", "us_jp_2y", "usdjpy", "audjpy"],
    }
    errors = sum(1 for feed in feed_health if feed["status"] == "error")
    stale = sum(1 for item in series.values() if item.get("stale"))
    base["feed_health"] = feed_health
    base["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    base["market_status"] = "All feeds current" if not errors and not stale else f"{errors} feed issue{'s' if errors != 1 else ''} · {stale} stale"
    base["build"] = os.getenv("GITHUB_SHA", "local")[:7]
    OUT.write_text(json.dumps(base, indent=2, ensure_ascii=False), encoding="utf-8")

    history = []
    if HIST.exists():
        try:
            history = json.loads(HIST.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    history.append({"generated_at": base["generated_at"], "series": {key: {"value": item.get("value"), "date": item.get("date")} for key, item in series.items()}})
    HIST.write_text(json.dumps(history[-370:], indent=2, ensure_ascii=False), encoding="utf-8")


if __name__ == "__main__":
    main()
